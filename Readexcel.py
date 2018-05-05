# -*- coding: utf-8 -*-
"""
Created on Wed May  2 22:56:19 2018

@author: User
"""
import pandas as pd
import itertools as itetls
from openpyxl import load_workbook
# Assign spreadsheet filename to `file`
file = 'execlFolder\\2018 LFS Sensor Data.xlsx'
wb=load_workbook(file)
#print(wb.get_sheet_by_name)
sheet = wb.get_sheet_by_name('01-02')
# Print the sheet title 
#print(sheet.title)
#

# Get currently active sheet
#anotherSheet = wb.active
#c=sheet['B2']
# Retrieve the value of a certain cell
#sheet['A1'].value
# Select element 'B2' of your sheet 
#c = sheet['B2']
# Retrieve the row number of your element
#c.row
# Retrieve the column letter of your element
#c.column
# Retrieve the coordinates of the cell 
#c.coordinate
# Check `anotherSheet` 
#print(anotherSheet)
# Retrieve cell value 
#print(sheet.cell(row=2, column=2).value)
# Print out values in column 2 
# =============================================================================
# for i in range(1, 4):
#      print(i, sheet.cell(row=i, column=3).value)
# =============================================================================
     # Print row per row
# =============================================================================
# print (sheet.max_row)
# print(sheet.max_column)
# for cellObj in sheet['A1':'C3']:
#       for cell in cellObj:
#               print(cell.coordinate, cell.value)
#       print('--- END ---')
#       
# =============================================================================
# Load spreadsheet

data = sheet.values
#print(list(data))
cols = next(data)[1:]

data=list(data)
idx=[r[0]for r in data]

data = (itetls.islice(r, 0, None,2) for r in data)

df=pd.DataFrame(data,columns=['時間','H1','H2','T1','T2'])
print(df['H2'])
# =============================================================================
# x1=pd.ExcelFile(file)
# xl = pd.read_excel(file,'01-02') 
# xl.head()
# # Print the sheet names
# print(x1.sheet_names[0])
# 
# 
# # Load a sheet into a DataFrame by name: df1
# df1 = x1.parse('01-01')
# print(df1)
# =============================================================================
